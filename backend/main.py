import json
import re
import shutil
import zlib
import sys
import xml.etree.ElementTree as ET
from io import BytesIO
from pathlib import Path
import os

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from git import Repo
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

global media_data_location


def get_session_with_retries() -> requests.Session:
    """Create a requests session with retry logic."""
    session = requests.Session()
    retries = Retry(
        total=5,
        backoff_factor=1,
        status_forcelist=[500, 502, 503, 504],
        allowed_methods=["GET"],
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session

# Pattern to parse =HYPERLINK("url","text") or =HYPERLINK("url", "text") formulas
HYPERLINK_PATTERN = re.compile(r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)')

def calculate_crc32(filepath: str) -> str:
    crc = 0
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            crc = zlib.crc32(chunk, crc)
    return format(crc & 0xFFFFFFFF, "08x")

def refresh_episode_metadata():
    repo_path = Path("data/eps-metadata")
    if (repo_path / ".git").exists():
        repo = Repo(repo_path)
        repo.remotes.origin.pull()
    else:
        Repo.clone_from("https://github.com/tissla/one-pace-jellyfin", repo_path)

def fetch_google_sheet_xlsx(sheet_id: str) -> dict[str, list[dict]]:
    """Fetch all sheets from a Google Sheet as XLSX to preserve hyperlinks."""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

    session = get_session_with_retries()
    response = session.get(url, timeout=300, stream=True)
    response.raise_for_status()

    # Download in chunks to handle large files better
    chunks = []
    for chunk in response.iter_content(chunk_size=8192):
        chunks.append(chunk)
    xlsx_data = BytesIO(b"".join(chunks))

    workbook = load_workbook(xlsx_data)
    all_sheets = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows())

        if not rows:
            all_sheets[sheet_name] = []
            continue

        # Get headers from first row
        headers = [cell.value for cell in rows[0]]

        data = []
        for row in rows[1:]:
            row_dict = {}
            has_data = False
            for header, cell in zip(headers, row):
                if header is None:
                    continue
                # Check if cell has a hyperlink object
                if cell.hyperlink:
                    row_dict[header] = {"text": cell.value, "link": cell.hyperlink.target}
                    has_data = True
                # Check if cell value is a =HYPERLINK() formula string
                elif isinstance(cell.value, str) and cell.value.startswith("=HYPERLINK"):
                    match = HYPERLINK_PATTERN.match(cell.value)
                    if match:
                        row_dict[header] = {"text": match.group(2), "link": match.group(1)}
                        has_data = True
                    else:
                        row_dict[header] = cell.value
                        has_data = True
                else:
                    row_dict[header] = cell.value
                    if cell.value is not None:
                        has_data = True
            # Skip rows where all values are null
            if has_data:
                data.append(row_dict)

        all_sheets[sheet_name] = data

    return all_sheets


def initialize_media(media_data_location: Path):
    """Copy episode metadata files from the cloned repo to the media data location."""
    source_dir = Path("data/eps-metadata/One Pace")

    if not source_dir.exists():
        refresh_episode_metadata()

    media_data_location.mkdir(parents=True, exist_ok=True)

    shutil.copytree(source_dir, media_data_location, dirs_exist_ok=True)

    print(f"Copied metadata from '{source_dir}' to '{media_data_location}'")


def refresh_onepace_sheet():
    """Download all sheets from the One Pace Google Sheet."""
    sheet_id = "1HQRMJgu_zArp-sLnvFMDzOyjdsht87eFLECxMK858lA"

    all_sheets = fetch_google_sheet_xlsx(sheet_id)

    output_dir = Path("data/sheets")
    output_dir.mkdir(parents=True, exist_ok=True)

    for sheet_name, rows in all_sheets.items():
        # Sanitize filename
        safe_name = sheet_name.replace("/", "-").replace(" ", "_").lower()
        output_path = output_dir / f"{safe_name}.json"

        with open(output_path, "w") as f:
            json.dump(rows, f, indent=2, default=str)

        print(f"Downloaded {len(rows)} rows to {output_path}")

    return all_sheets


def build_season_to_arc_map(arc_overview: list[dict]) -> dict[int, dict]:
    """
    Build a mapping from integer season numbers to arc information.

    The NFO files use integer seasons (1-36), but arc_overview.json uses
    fractional arc numbers (1.0, 6.5, 9.5, etc.). This function sorts arcs
    by their number and assigns sequential integer season numbers.

    Args:
        arc_overview: List of arc data from arc_overview.json

    Returns:
        Dict mapping season number to arc info with keys:
        - arc_name: Cleaned arc name (without TBR/WIP suffixes)
        - arc_no: Original arc number from sheets
        - json_filename: Expected JSON filename for this arc's episodes
    """
    # Filter out rows with null or non-numeric "No." (totals/notes rows)
    valid_arcs = [
        arc for arc in arc_overview
        if arc.get("No.") is not None and isinstance(arc.get("No."), (int, float))
    ]

    # Sort by arc number
    sorted_arcs = sorted(valid_arcs, key=lambda x: float(x["No."]))

    season_map = {}
    for season_num, arc in enumerate(sorted_arcs, start=1):
        # Handle both dict format {"text": "...", "link": ...} and plain string
        arc_name_raw = arc["Arcs"]["text"] if isinstance(arc["Arcs"], dict) else arc["Arcs"]

        # Clean arc name: remove (TBR), (WIP), extra spaces
        arc_name_clean = re.sub(r'\s*\((TBR|WIP)\)\s*', '', arc_name_raw).strip()

        # Generate JSON filename matching refresh_onepace_sheet() logic
        # Note: sheet names may differ slightly from arc overview names
        json_filename = (
            arc_name_clean
            .replace("/", "-")
            .replace(" ", "_")
            .replace("'", "")  # Apostrophes are stripped
            .lower()
            + ".json"
        )

        season_map[season_num] = {
            "arc_name": arc_name_clean,
            "arc_no": arc["No."],
            "json_filename": json_filename
        }

    return season_map


def parse_episode_number(ep_value: str) -> int:
    """
    Extract episode number from "Arc Name ##" format.

    Args:
        ep_value: Episode identifier like "Romance Dawn 01" or "Egghead 08"

    Returns:
        Episode number as integer, or 1 for single-episode arcs without number
    """
    ep_value = ep_value.strip()

    # Try to match trailing digits (handles "01", "08", "00", etc.)
    match = re.search(r'(\d+)$', ep_value)
    if match:
        return int(match.group(1))

    # Single-episode arc (no number suffix)
    return 1


def find_arc_json_file(sheets_dir: Path, json_filename: str) -> Path | None:
    """
    Find the JSON file for an arc, handling naming mismatches.

    Some arc names in arc_overview.json don't exactly match the sheet tab names,
    e.g., "The Adventures of the Straw Hats" vs "The Adventures of the Straw Hat".
    """
    json_path = sheets_dir / json_filename
    if json_path.exists():
        return json_path

    # Try common variations
    base_name = json_filename[:-5]  # Remove .json

    # Try removing trailing 's' (Straw Hats -> Straw Hat)
    if base_name.endswith("s"):
        alt_path = sheets_dir / f"{base_name[:-1]}.json"
        if alt_path.exists():
            return alt_path

    # Try adding trailing 's'
    alt_path = sheets_dir / f"{base_name}s.json"
    if alt_path.exists():
        return alt_path

    return None


def load_arc_episodes(sheets_dir: Path, season_map: dict[int, dict]) -> dict[tuple[str, int], dict]:
    """
    Load all arc JSON files and create a mapping from (arc_name, ep_num) to row data.

    Args:
        sheets_dir: Path to the sheets directory containing arc JSON files
        season_map: Mapping from season number to arc info

    Returns:
        Dict mapping (arc_name, episode_number) tuples to sheet row data
    """
    arc_episode_map = {}

    for arc_info in season_map.values():
        json_path = find_arc_json_file(sheets_dir, arc_info["json_filename"])

        if json_path is None:
            print(f"Warning: Missing JSON file for {arc_info['arc_name']}: {arc_info['json_filename']}")
            continue

        with open(json_path) as f:
            episodes = json.load(f)

        for row in episodes:
            # Find the "One Pace Episode" column (may have leading space)
            ep_col_value = None
            for key in row:
                if key is not None and key.strip() == "One Pace Episode":
                    ep_col_value = row[key]
                    break

            if ep_col_value is None or not isinstance(ep_col_value, str):
                continue

            ep_num = parse_episode_number(ep_col_value)
            arc_episode_map[(arc_info["arc_name"], ep_num)] = row

    return arc_episode_map


def parse_nfo_files(metadata_dir: Path) -> list[dict]:
    """
    Parse all episode NFO files and extract metadata.

    Args:
        metadata_dir: Path to the "One Pace" metadata directory

    Returns:
        List of dicts with keys: filename, season, episode, title
    """
    episodes = []

    for season_dir in sorted(metadata_dir.glob("Season *")):
        for nfo_file in sorted(season_dir.glob("One Pace - S*E* - *.nfo")):
            try:
                tree = ET.parse(nfo_file)
                root = tree.getroot()

                episodes.append({
                    "filename": nfo_file.stem,  # Without .nfo extension
                    "season": int(root.findtext("season")),
                    "episode": int(root.findtext("episode")),
                    "title": root.findtext("title"),
                })
            except Exception as e:
                print(f"Error parsing {nfo_file}: {e}")

    return episodes


def build_episode_mapping() -> list[dict]:
    """
    Build a complete mapping of all One Pace episodes with metadata from NFO files
    and torrent information from Google Sheets.

    This function:
    1. Parses arc_overview.json to map season numbers to arc names
    2. Loads each arc's JSON file to get torrent links and CRC32 checksums
    3. Parses NFO files to get episode metadata (title, season, episode number)
    4. Joins the data to create a unified episode mapping

    Returns:
        List of dictionaries, each containing:
        - ep_name: str - Episode name (NFO filename without extension)
        - season: int - Season number
        - ep_number: int - Episode number within season
        - file_location_media: str - Relative path to media file
        - torrent_link: str | None - Nyaa torrent link
        - crc32: str | None - CRC32 checksum of the MKV file
        - torrent_link_extended: str | None - Extended version torrent link
        - crc32_extended: str | None - Extended version CRC32 checksum
    """
    metadata_dir = Path("data/eps-metadata/One Pace")
    sheets_dir = Path("data/sheets")

    # Ensure data is available
    if not metadata_dir.exists():
        refresh_episode_metadata()
    if not sheets_dir.exists():
        refresh_onepace_sheet()

    # Step 1: Load and parse arc overview
    with open(sheets_dir / "arc_overview.json") as f:
        arc_overview = json.load(f)

    season_map = build_season_to_arc_map(arc_overview)

    # Step 2: Load arc episode data
    arc_episode_map = load_arc_episodes(sheets_dir, season_map)

    # Step 3: Parse NFO files
    nfo_episodes = parse_nfo_files(metadata_dir)

    # Step 4: Build mappings
    results = []

    for nfo in nfo_episodes:
        season_num = nfo["season"]
        ep_num = nfo["episode"]

        # Get arc info for this season
        arc_info = season_map.get(season_num)
        if arc_info is None:
            print(f"Warning: No arc mapping for season {season_num}")
            continue

        # Look up sheet data
        sheet_key = (arc_info["arc_name"], ep_num)
        sheet_row = arc_episode_map.get(sheet_key)

        # Extract torrent info from sheet row
        torrent_link = None
        crc32 = None
        torrent_link_extended = None
        crc32_extended = None

        if sheet_row:
            mkv_crc32 = sheet_row.get("MKV CRC32")
            if isinstance(mkv_crc32, dict):
                torrent_link = mkv_crc32.get("link")
                crc32 = mkv_crc32.get("text")

            mkv_crc32_ext = sheet_row.get("MKV CRC32 (Extended)")
            if isinstance(mkv_crc32_ext, dict):
                torrent_link_extended = mkv_crc32_ext.get("link")
                crc32_extended = mkv_crc32_ext.get("text")

        # Build file location path
        file_location = f"Season {season_num}/{nfo['filename']}.mkv"

        results.append({
            "ep_name": nfo["filename"],
            "season": season_num,
            "ep_number": ep_num,
            "file_location_media": file_location,
            "torrent_link": torrent_link,
            "crc32": crc32,
            "torrent_link_extended": torrent_link_extended,
            "crc32_extended": crc32_extended,
        })

    return results


if __name__ == "__main__":
    # if len(sys.argv) != 2:
    #     print(f"Usage: {sys.argv[0]} <filepath>")
    #     sys.exit(1)

    # filepath = sys.argv[1]
    # checksum = calculate_crc32(filepath)
    # print(f"CRC32: {checksum}")
    media_data_location = Path(os.getenv("MEDIA_DATA_LOCATION", "data/media"))
    # initialize_media(media_data_location)
    metadata_mapping = build_episode_mapping()
    print(f"Built metadata mapping for {len(metadata_mapping)} episodes.")
    print(metadata_mapping)
    # refresh_episode_metadata()
    # refresh_onepace_sheet()