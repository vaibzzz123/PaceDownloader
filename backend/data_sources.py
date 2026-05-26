"""Data source functions for downloading raw data and storing it in the data directory."""

import json
import re
from email.utils import parsedate_to_datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from git import Repo
from openpyxl import load_workbook

from logging_config import get_logger

logger = get_logger(__name__)

GITHUB_REPO_URL = "https://github.com/tissla/one-pace-jellyfin"
GOOGLE_SHEET_ID = "1HQRMJgu_zArp-sLnvFMDzOyjdsht87eFLECxMK858lA"
ONEPACE_RELEASES_RSS_URL = "https://onepace.net/en/releases/rss.xml"

METADATA_DIR = Path("data/eps-metadata")
SHEETS_DIR = Path("data/sheets")
RELEASES_DIR = Path("data/releases")
RELEASES_JSON_PATH = RELEASES_DIR / "onepace_releases.json"

RSS_NS = {
    "rss": "https://www.rssboard.org/rss-specification",
    "torrent": "http://xmlns.ezrss.it/0.1/",
}


def _get_session_with_retries() -> requests.Session:
    """Create a requests session with retry logic."""
    session = requests.Session()
    retries = Retry(
        total=5,
        backoff_factor=1,
        status_forcelist=[500, 502, 503, 504],
        allowed_methods=["GET"],
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def _normalize_release_title(title: str | None) -> str:
    """Normalize release titles for later lookup without changing displayed text."""
    if not title:
        return ""
    normalized = re.sub(r"\s+", " ", title).strip().lower()
    return re.sub(r"[^a-z0-9]+", " ", normalized).strip()


def _extract_nyaa_id(*urls: str | None) -> int | None:
    """Extract a nyaa.si numeric ID from known view/download URL shapes."""
    for url in urls:
        if not url:
            continue
        match = re.search(r"nyaa\.si/(?:view|download)/(\d+)", url)
        if match:
            return int(match.group(1))
    return None


def _parse_publication_date(pub_date: str | None) -> str | None:
    """Return the ISO date parsed from an RSS pubDate value."""
    if not pub_date:
        return None
    parsed = parsedate_to_datetime(pub_date)
    return parsed.date().isoformat()


def parse_onepace_releases_rss(xml_text: str) -> list[dict]:
    """Parse One Pace releases RSS into JSON-serializable release records."""
    root = ET.fromstring(xml_text)
    releases = []

    for item in root.findall(".//rss:item", RSS_NS):
        title = item.findtext("rss:title", default="", namespaces=RSS_NS).strip()
        pub_date_raw = item.findtext("rss:pubDate", default="", namespaces=RSS_NS).strip()
        publication_date = _parse_publication_date(pub_date_raw)
        categories = [
            category.text.strip()
            for category in item.findall("rss:category", RSS_NS)
            if category.text and category.text.strip()
        ]

        link = item.findtext("rss:link", default="", namespaces=RSS_NS).strip() or None
        enclosure = item.find("rss:enclosure", RSS_NS)
        torrent_url = enclosure.get("url") if enclosure is not None else None
        info_hash = item.findtext("torrent:infoHash", default="", namespaces=RSS_NS).strip()
        magnet_uri = item.findtext("torrent:magnetURI", default="", namespaces=RSS_NS).strip()
        torrent_file_name = item.findtext("torrent:fileName", default="", namespaces=RSS_NS).strip()

        nyaa_url = link if link and "nyaa.si" in link else None
        nyaa_id = _extract_nyaa_id(nyaa_url, torrent_url)

        releases.append({
            "title": title,
            "normalized_title": _normalize_release_title(title),
            "publication_date": publication_date,
            "categories": categories,
            "nyaa_url": nyaa_url,
            "nyaa_id": nyaa_id,
            "torrent_url": torrent_url,
            "magnet_uri": magnet_uri,
            "info_hash": info_hash.lower() if info_hash else None,
            "torrent_file_name": torrent_file_name,
        })

    return releases


def _metadata_dir_contains_only_empty_static_placeholder() -> bool:
    """Return whether only the poster StaticFiles placeholder exists."""
    if not METADATA_DIR.exists() or (METADATA_DIR / ".git").exists():
        return False

    entries = list(METADATA_DIR.iterdir())
    placeholder_dir = METADATA_DIR / "One Pace"
    return (
        len(entries) == 1
        and entries[0] == placeholder_dir
        and placeholder_dir.is_dir()
        and not any(placeholder_dir.iterdir())
    )


def _prepare_metadata_clone_target() -> None:
    if not METADATA_DIR.exists() or (METADATA_DIR / ".git").exists():
        return

    if _metadata_dir_contains_only_empty_static_placeholder():
        logger.info("Removing empty metadata placeholder before initial clone")
        (METADATA_DIR / "One Pace").rmdir()
        return

    if any(METADATA_DIR.iterdir()):
        raise RuntimeError(
            f"Cannot clone episode metadata into non-git directory with existing files: {METADATA_DIR}"
        )


def fetch_episode_metadata():
    """Clone or pull the One Pace Jellyfin metadata repository."""
    if (METADATA_DIR / ".git").exists():
        repo = Repo(METADATA_DIR)
        logger.info("Pulling latest episode metadata")
        repo.remotes.origin.pull()
        logger.debug("Git pull completed for %s", METADATA_DIR)
    else:
        logger.info("Cloning One Pace Jellyfin metadata repository")
        _prepare_metadata_clone_target()
        Repo.clone_from(GITHUB_REPO_URL, METADATA_DIR)
        logger.debug("Git clone completed to %s", METADATA_DIR)


def _fetch_google_sheet_xlsx(
    sheet_id: str, save_xlsx: bool = False
) -> dict[str, list[dict]]:
    """Fetch all sheets from a Google Sheet as XLSX to preserve hyperlinks.

    Args:
        sheet_id: The Google Sheet ID to fetch
        save_xlsx: If True, save a copy of the downloaded XLSX file for debugging
    """
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    logger.debug("Fetching Google Sheet: %s", sheet_id)

    SHEETS_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = SHEETS_DIR / "onepace_sheets.xlsx"

    session = _get_session_with_retries()
    max_attempts = 3
    for attempt in range(1, max_attempts + 1):
        try:
            response = session.get(url, timeout=300, stream=True)
            response.raise_for_status()
            logger.debug("Google Sheets response status: %d", response.status_code)
            with open(xlsx_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=65536):
                    f.write(chunk)
            break
        except Exception as e:
            xlsx_path.unlink(missing_ok=True)
            if attempt < max_attempts:
                logger.warning("Attempt %d/%d failed fetching Google Sheet: %s", attempt, max_attempts, e)
            else:
                logger.error("Failed to fetch Google Sheet after %d attempts: %s", max_attempts, e)
                raise

    try:
        workbook = load_workbook(xlsx_path)
        all_sheets = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows())

            if not rows:
                all_sheets[sheet_name] = []
                continue

            # Get headers from first row
            headers = [cell.value for cell in rows[0]]

            data = []
            for row in rows[1:]:
                row_dict = {}
                has_data = False
                for header, cell in zip(headers, row):
                    if header is None:
                        continue
                    # Check if cell has a hyperlink object
                    if cell.hyperlink:
                        row_dict[header] = {
                            "text": cell.value,
                            "link": cell.hyperlink.target,
                        }
                        has_data = True
                    # Check if cell value is a =HYPERLINK() formula string
                    elif isinstance(cell.value, str) and cell.value.startswith(
                        "=HYPERLINK"
                    ):
                        # Pattern to parse =HYPERLINK("url","text") or =HYPERLINK("url", "text") formulas
                        HYPERLINK_PATTERN = re.compile(
                            r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)'
                        )
                        match = HYPERLINK_PATTERN.match(cell.value)
                        if match:
                            row_dict[header] = {
                                "text": match.group(2),
                                "link": match.group(1),
                            }
                            has_data = True
                        else:
                            row_dict[header] = cell.value
                            has_data = True
                    else:
                        row_dict[header] = cell.value
                        if cell.value is not None:
                            has_data = True
                # Skip rows where all values are null
                if has_data:
                    data.append(row_dict)

            all_sheets[sheet_name] = data

        return all_sheets
    finally:
        if not save_xlsx:
            xlsx_path.unlink(missing_ok=True)


def fetch_onepace_sheet():
    """Download the One Pace Google Sheet and save each tab as JSON."""
    all_sheets = _fetch_google_sheet_xlsx(GOOGLE_SHEET_ID)

    SHEETS_DIR.mkdir(parents=True, exist_ok=True)

    for sheet_name, rows in all_sheets.items():
        safe_name = sheet_name.replace("/", "-").replace(" ", "_").lower()
        output_path = SHEETS_DIR / f"{safe_name}.json"

        with open(output_path, "w") as f:
            json.dump(rows, f, indent=2, default=str)

        logger.info("Saved %d rows to %s", len(rows), output_path)


def fetch_onepace_releases():
    """Download the One Pace releases RSS feed and save parsed releases as JSON."""
    logger.debug("Fetching One Pace releases RSS: %s", ONEPACE_RELEASES_RSS_URL)

    session = _get_session_with_retries()
    response = session.get(ONEPACE_RELEASES_RSS_URL, timeout=60)
    response.raise_for_status()

    releases = parse_onepace_releases_rss(response.text)

    RELEASES_DIR.mkdir(parents=True, exist_ok=True)
    with open(RELEASES_JSON_PATH, "w") as f:
        json.dump(releases, f, indent=2)

    logger.info("Saved %d One Pace releases to %s", len(releases), RELEASES_JSON_PATH)
